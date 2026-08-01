"""Offline, staged physical roller inventory knowledge-base operations.

This module intentionally stops at design/asset inventory and reviewable candidate
data. It does not resolve a drawing occurrence to a physical asset automatically.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from hashlib import sha256
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

from rollform_extractor.database import (
    RollerAlias,
    RollerAsset,
    RollerAuditEvent,
    RollerDesign,
    RollerGeometryRevision,
    RollerImportBatch,
    RollerImportRow,
    RollerLocation,
    RollerRegrindHistory,
)

INVENTORY_SCHEMA_VERSION = 1
UNKNOWN_VALUES = {"", "-", "na", "n/a", "none", "null", "unknown", "tbd", "?"}
REQUIRED_FIELDS = ("design_id", "asset_id")
TEMPLATE_FIELDS = (
    "design_id", "design_name", "design_type", "manufacturer", "design_alias",
    "asset_id", "serial_number", "condition", "location_id", "location_name",
    "revision_id", "diameter", "diameter_unit", "bore", "bore_unit",
    "width", "width_unit", "measurement_method", "geometry_source",
    "verification_status", "confidence", "physical_fingerprint", "shape_fingerprint", "source_file", "notes",
)


@dataclass(frozen=True)
class InventoryRow:
    row_number: int
    original: dict[str, Any]
    normalized: dict[str, Any]


@dataclass(frozen=True)
class InventoryValidation:
    source_name: str
    source_sha256: str
    row_count: int
    accepted_count: int
    review_count: int
    rejected_count: int
    rows: tuple[InventoryRow, ...] = ()
    errors: tuple[str, ...] = ()
    duplicate_candidates: tuple[dict[str, Any], ...] = ()

    @property
    def valid(self) -> bool:
        return not self.errors and self.rejected_count == 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": INVENTORY_SCHEMA_VERSION,
            "source_name": self.source_name,
            "source_sha256": self.source_sha256,
            "row_count": self.row_count,
            "accepted_count": self.accepted_count,
            "review_count": self.review_count,
            "rejected_count": self.rejected_count,
            "valid": self.valid,
            "errors": list(self.errors),
            "duplicate_candidates": list(self.duplicate_candidates),
            "rows": [
                {"row_number": row.row_number, "original": row.original, "normalized": row.normalized}
                for row in self.rows
            ],
        }


@dataclass(frozen=True)
class InventoryImportSummary:
    batch_id: int
    status: str
    source_sha256: str
    accepted: int
    review_required: int
    rejected: int
    idempotent: bool = False
    errors: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, Any]:
        return {
            "batch_id": self.batch_id,
            "status": self.status,
            "source_sha256": self.source_sha256,
            "accepted": self.accepted,
            "review_required": self.review_required,
            "rejected": self.rejected,
            "idempotent": self.idempotent,
            "errors": list(self.errors),
        }


def file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    aliases = {"id": "asset_id", "roller_id": "asset_id", "design": "design_id", "alias": "design_alias"}
    return aliases.get(text, text)


def normalize_value(value: object) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if text.lower() in UNKNOWN_VALUES:
        return None
    try:
        return float(text) if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text) else text
    except ValueError:
        return text


def _read_rows(path: Path) -> list[InventoryRow]:
    suffix = path.suffix.lower()
    raw_rows: list[tuple[int, dict[str, Any]]] = []
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            raw_rows = [(i, dict(row)) for i, row in enumerate(csv.DictReader(handle), start=2)]
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=True, read_only=True)
        values = workbook.active.iter_rows(values_only=True)
        headers = [normalize_header(value) for value in next(values, ())]
        raw_rows = [(i, dict(zip(headers, row, strict=False))) for i, row in enumerate(values, start=2) if any(row)]
    else:
        raise ValueError(f"unsupported inventory file: {path.suffix}; expected .csv or .xlsx")
    rows: list[InventoryRow] = []
    for row_number, original in raw_rows:
        original = {normalize_header(key): value for key, value in original.items() if normalize_header(key)}
        normalized = {key: normalize_value(value) for key, value in original.items()}
        rows.append(InventoryRow(row_number, original, normalized))
    return rows


def _unit_status(row: dict[str, Any]) -> str:
    units = [row.get(f"{name}_unit") for name in ("diameter", "bore", "width") if row.get(name) is not None]
    if not units or any(unit is None for unit in units):
        return "UNKNOWN"
    normalized = {str(unit).strip().lower() for unit in units}
    known = {"mm", "millimetre", "millimeter", "in", "inch", "cm", "m"}
    if not normalized.issubset(known):
        return "UNKNOWN"
    return "CONFIRMED" if len(normalized) == 1 else "MIXED"


def _dimension_values(row: dict[str, Any]) -> dict[str, Any]:
    dimensions: dict[str, Any] = {}
    for name in ("diameter", "bore", "width"):
        value = row.get(name)
        unit = row.get(f"{name}_unit")
        millimetres = None
        conversion = {"mm": 1.0, "millimetre": 1.0, "millimeter": 1.0, "in": 25.4, "inch": 25.4, "cm": 10.0, "m": 1000.0}.get(str(unit or "").lower())
        if isinstance(value, (int, float)) and conversion is not None and row.get("verification_status") == "VERIFIED":
            millimetres = float(value) * conversion
        dimensions[name] = {"original_value": value, "original_unit": unit, "millimetres": millimetres}
    dimensions["conversion_factor_to_mm"] = {name: (item["millimetres"] / item["original_value"] if item["millimetres"] is not None and item["original_value"] else None) for name, item in dimensions.items() if isinstance(item, dict)}
    return dimensions


def _physical_fingerprint(dimensions: dict[str, Any]) -> str | None:
    values = {name: item.get("millimetres") for name, item in dimensions.items() if isinstance(item, dict) and "millimetres" in item}
    if not values or any(value is None for value in values.values()):
        return None
    return sha256(json.dumps(values, sort_keys=True, separators=(",", ":")).encode()).hexdigest()


def validate_inventory(path: Path, engine: Engine | None = None) -> InventoryValidation:
    rows = _read_rows(path)
    source_hash = file_sha256(path)
    errors: list[str] = []
    duplicate_candidates: list[dict[str, Any]] = []
    seen_designs: dict[str, int] = {}
    seen_assets: dict[str, int] = {}
    seen_aliases: dict[str, int] = {}
    accepted = review = rejected = 0
    validated: list[InventoryRow] = []
    existing_designs: set[str] = set()
    existing_assets: set[str] = set()
    existing_aliases: set[str] = set()
    if engine is not None:
        with Session(engine) as session:
            existing_designs = {value for value in session.scalars(select(RollerDesign.design_id))}
            existing_assets = {value for value in session.scalars(select(RollerAsset.asset_id))}
            existing_aliases = {value for value in session.scalars(select(RollerAlias.normalized_alias))}
    for row in rows:
        data = row.normalized
        reasons: list[str] = []
        for field_name in REQUIRED_FIELDS:
            if not data.get(field_name):
                reasons.append(f"MISSING_{field_name.upper()}")
        design_id = str(data.get("design_id") or "")
        asset_id = str(data.get("asset_id") or "")
        alias = str(data.get("design_alias") or "").strip().lower()
        if asset_id in seen_assets:
            reasons.append("DUPLICATE_ASSET_ID")
        if alias and (alias in seen_aliases or alias in existing_aliases):
            reasons.append("ALIAS_COLLISION")
        if asset_id in existing_assets:
            duplicate_candidates.append({"row_number": row.row_number, "kind": "asset", "key": asset_id})
        verification = str(data.get("verification_status") or "UNVERIFIED").upper()
        if verification == "VERIFIED" and _unit_status(data) != "CONFIRMED":
            reasons.append("UNKNOWN_UNITS_BLOCK_VERIFICATION")
        if reasons:
            if any(reason.startswith("DUPLICATE") or reason.endswith("COLLISION") for reason in reasons):
                review += 1
            else:
                rejected += 1
        else:
            accepted += 1
        validated.append(row)
        if design_id:
            seen_designs[design_id] = row.row_number
        if asset_id:
            seen_assets[asset_id] = row.row_number
        if alias:
            seen_aliases[alias] = row.row_number
    return InventoryValidation(path.name, source_hash, len(rows), accepted, review, rejected, tuple(validated), tuple(errors), tuple(duplicate_candidates))


def import_inventory(path: Path, engine: Engine, *, actor: str = "inventory_import") -> InventoryImportSummary:
    validation = validate_inventory(path, engine)
    with Session(engine) as session:
        existing = session.scalar(select(RollerImportBatch).where(RollerImportBatch.source_sha256 == validation.source_sha256))
        if existing is not None:
            return InventoryImportSummary(existing.id, existing.status, validation.source_sha256, existing.accepted_count, existing.review_count, existing.rejected_count, True)
    row_reasons: dict[int, list[str]] = {}
    seen_designs: set[str] = set()
    seen_assets: set[str] = set()
    seen_aliases: set[str] = set()
    for row in validation.rows:
        data = row.normalized
        reasons: list[str] = []
        design_id = str(data.get("design_id") or "")
        asset_id = str(data.get("asset_id") or "")
        alias = str(data.get("design_alias") or "").strip().lower()
        if not design_id:
            reasons.append("MISSING_DESIGN_ID")
        if not asset_id:
            reasons.append("MISSING_ASSET_ID")
        if asset_id in seen_assets or (alias and alias in seen_aliases):
            reasons.append("DUPLICATE_CANDIDATE_REQUIRES_REVIEW")
        if str(data.get("verification_status") or "UNVERIFIED").upper() == "VERIFIED" and _unit_status(data) != "CONFIRMED":
            reasons.append("UNKNOWN_UNITS_BLOCK_VERIFICATION")
        if any(item["row_number"] == row.row_number for item in validation.duplicate_candidates):
            reasons.append("DUPLICATE_CANDIDATE_REQUIRES_REVIEW")
        row_reasons[row.row_number] = reasons
        if design_id:
            seen_designs.add(design_id)
        if asset_id:
            seen_assets.add(asset_id)
        if alias:
            seen_aliases.add(alias)
    with Session(engine) as session, session.begin():
        batch = RollerImportBatch(source_name=path.name, source_sha256=validation.source_sha256, source_path=str(path), row_count=len(validation.rows))
        session.add(batch)
        session.flush()
        accepted = review = rejected = 0
        for row in validation.rows:
            reasons = row_reasons[row.row_number]
            status = "ACCEPTED"
            if reasons:
                status = "REVIEW_REQUIRED" if "DUPLICATE_CANDIDATE_REQUIRES_REVIEW" in reasons else "REJECTED"
            if status == "ACCEPTED":
                design_id = str(row.normalized["design_id"])
                location_id = row.normalized.get("location_id")
                if location_id and session.get(RollerLocation, str(location_id)) is None:
                    session.add(RollerLocation(location_id=str(location_id), name=str(row.normalized.get("location_name") or location_id), location_type="INVENTORY", provenance_json={"batch_id": batch.id, "source_row": row.row_number}))
                design = session.get(RollerDesign, design_id)
                if design is None:
                    design = RollerDesign(design_id=design_id, name=row.normalized.get("design_name"), design_type=row.normalized.get("design_type"), manufacturer=row.normalized.get("manufacturer"), status="CANDIDATE", verified=0, provenance_json={"batch_id": batch.id, "source_row": row.row_number})
                    session.add(design)
                session.flush()
                asset = RollerAsset(asset_id=str(row.normalized["asset_id"]), design_id=design_id, serial_number=row.normalized.get("serial_number"), condition=row.normalized.get("condition"), location_id=str(location_id) if location_id else None, verified=0, source=str(path), provenance_json={"batch_id": batch.id, "source_row": row.row_number})
                session.add(asset)
                session.flush()
                if row.normalized.get("design_alias"):
                    alias = str(row.normalized["design_alias"])
                    session.add(RollerAlias(design_id=design_id, alias=alias, normalized_alias=alias.strip().lower(), source=str(path), verified=0, provenance_json={"batch_id": batch.id, "source_row": row.row_number}))
                revision_id = row.normalized.get("revision_id")
                if revision_id:
                    session.flush()
                    dimensions = _dimension_values(row.normalized)
                    session.add(RollerGeometryRevision(revision_id=str(revision_id), design_id=design_id, asset_id=asset.id, dimensions_json=dimensions, unit_status=_unit_status(row.normalized), measurement_method=row.normalized.get("measurement_method"), source=row.normalized.get("geometry_source") or str(path), confidence=_number(row.normalized.get("confidence")), verification_status=str(row.normalized.get("verification_status") or "UNVERIFIED").upper(), input_file_hash=validation.source_sha256, algorithm_version="phase16-v1", configuration_hash=f"inventory-schema-{INVENTORY_SCHEMA_VERSION}", physical_fingerprint=row.normalized.get("physical_fingerprint") or _physical_fingerprint(dimensions), shape_fingerprint=row.normalized.get("shape_fingerprint"), provenance_json={"batch_id": batch.id, "source_row": row.row_number}))
                accepted += 1
            elif status == "REVIEW_REQUIRED":
                review += 1
            else:
                rejected += 1
            session.add(RollerImportRow(batch_id=batch.id, row_number=row.row_number, original_json=row.original, normalized_json=row.normalized, status=status, reasons_json=reasons))
        batch.accepted_count, batch.review_count, batch.rejected_count = accepted, review, rejected
        batch.status = "IMPORTED_WITH_REVIEW" if review else ("IMPORTED" if rejected == 0 else "IMPORTED_WITH_REJECTIONS")
        session.add(RollerAuditEvent(entity_type="roller_import_batch", entity_key=str(batch.id), action="IMPORT", actor=actor, before_json=None, after_json={"status": batch.status, "accepted": accepted, "review": review, "rejected": rejected}, source=str(path)))
        return InventoryImportSummary(batch.id, batch.status, validation.source_sha256, accepted, review, rejected)


def _number(value: Any) -> float | None:
    try:
        return float(value) if value is not None else None
    except (TypeError, ValueError):
        return None


def inventory_stats(engine: Engine) -> dict[str, int]:
    with Session(engine) as session:
        return {"designs": session.scalar(select(func.count()).select_from(RollerDesign)) or 0, "assets": session.scalar(select(func.count()).select_from(RollerAsset)) or 0, "geometry_revisions": session.scalar(select(func.count()).select_from(RollerGeometryRevision)) or 0, "aliases": session.scalar(select(func.count()).select_from(RollerAlias)) or 0, "import_batches": session.scalar(select(func.count()).select_from(RollerImportBatch)) or 0, "review_rows": session.scalar(select(func.count()).select_from(RollerImportRow).where(RollerImportRow.status == "REVIEW_REQUIRED")) or 0}


def export_inventory(engine: Engine, output: Path) -> Path:
    output.mkdir(parents=True, exist_ok=True)
    with Session(engine) as session:
        rows = []
        for asset in session.scalars(select(RollerAsset).order_by(RollerAsset.asset_id)):
            design = session.get(RollerDesign, asset.design_id) if asset.design_id else None
            rows.append({"design_id": asset.design_id, "design_name": design.name if design else None, "asset_id": asset.asset_id, "serial_number": asset.serial_number, "condition": asset.condition, "location_id": asset.location_id, "verified": bool(asset.verified)})
    target = output / "roller_inventory.csv"
    with target.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]) if rows else ["design_id", "asset_id"])
        writer.writeheader()
        writer.writerows(rows)
    return target


def export_rejected_rows(engine: Engine, batch_id: int, output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    with Session(engine) as session:
        rows = session.scalars(select(RollerImportRow).where(RollerImportRow.batch_id == batch_id, RollerImportRow.status != "ACCEPTED").order_by(RollerImportRow.row_number)).all()
    with output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["row_number", "status", "reasons", "original_json"])
        writer.writeheader()
        for row in rows:
            writer.writerow({"row_number": row.row_number, "status": row.status, "reasons": ";".join(row.reasons_json), "original_json": json.dumps(row.original_json, sort_keys=True)})
    return output


def write_inventory_template(target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.suffix.lower() == ".xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(list(TEMPLATE_FIELDS))
        workbook.save(target)
    else:
        with target.open("w", newline="", encoding="utf-8") as handle:
            csv.writer(handle).writerow(TEMPLATE_FIELDS)
    return target
