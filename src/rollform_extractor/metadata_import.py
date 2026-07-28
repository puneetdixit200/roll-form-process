from __future__ import annotations

import csv
from dataclasses import dataclass
import json
from pathlib import Path
import re
import sqlite3

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

from rollform_extractor.database import Project, ProjectCode, ProjectMetadata, ResultProvenance


APPROVED_FIELDS = (
    "material_grade",
    "steel_thickness",
    "strip_width",
    "coil_width",
    "machine_id",
    "shaft_diameter",
    "product_code",
    "customer_code",
    "production_status",
    "tooling_worked",
    "known_defects",
    "engineer_notes",
    "copra_project_reference",
)
UNKNOWN_VALUES = {"", "-", "na", "n/a", "none", "null", "unknown"}
PROJECT_CODE_RE = re.compile(r"\b[A-Z]\d{4}\b", re.IGNORECASE)


@dataclass(frozen=True)
class ProjectCodeResolution:
    drawing_id: str
    related_project_codes: tuple[str, ...]


@dataclass(frozen=True)
class MetadataImportSummary:
    imported: int
    unmatched: tuple[str, ...] = ()
    conflicts: tuple[str, ...] = ()


def resolve_project_codes(filename: str) -> ProjectCodeResolution:
    drawing_id = Path(filename).stem
    codes = tuple(dict.fromkeys(code.upper() for code in PROJECT_CODE_RE.findall(drawing_id)))
    return ProjectCodeResolution(drawing_id=drawing_id, related_project_codes=codes)


def import_metadata(path: Path, engine: Engine | sqlite3.Connection) -> MetadataImportSummary:
    rows = _read_rows(path)
    if isinstance(engine, sqlite3.Connection):
        return _import_metadata_sqlite(path, engine, rows)
    imported = 0
    unmatched: list[str] = []
    conflicts: list[str] = []
    with Session(engine) as session, session.begin():
        for row_number, row in rows:
            project = _resolve_row_project(session, row)
            row_name = row.get("drawing_id") or row.get("project_code") or "<missing>"
            if project is None:
                unmatched.append(f"{row_number}:{row_name}")
                continue
            row_conflicts = _upsert_metadata(session, project.id, path, row_number, row)
            if row_conflicts:
                conflicts.extend(f"{row_number}:{project.drawing_id}:{field}" for field in row_conflicts)
                continue
            imported += 1
            _ensure_project_codes(session, project, row)
    return MetadataImportSummary(imported=imported, unmatched=tuple(unmatched), conflicts=tuple(conflicts))


def _import_metadata_sqlite(
    path: Path,
    db: sqlite3.Connection,
    rows: list[tuple[int, dict[str, str | None]]],
) -> MetadataImportSummary:
    _ensure_sqlite_metadata_schema(db)
    imported = 0
    unmatched: list[str] = []
    conflicts: list[str] = []
    for row_number, row in rows:
        project = _resolve_sqlite_project(db, row)
        row_name = row.get("drawing_id") or row.get("project_code") or "<missing>"
        if project is None:
            unmatched.append(f"{row_number}:{row_name}")
            continue
        project_id, drawing_id = project
        row_conflicts = _upsert_sqlite_metadata(db, project_id, path, row_number, row)
        if row_conflicts:
            conflicts.extend(f"{row_number}:{drawing_id}:{field}" for field in row_conflicts)
            continue
        imported += 1
        _ensure_sqlite_project_codes(db, project_id, drawing_id, row)
    db.commit()
    return MetadataImportSummary(imported=imported, unmatched=tuple(unmatched), conflicts=tuple(conflicts))


def _ensure_sqlite_metadata_schema(db: sqlite3.Connection) -> None:
    db.executescript(
        """
        create table if not exists project_codes (
            code text primary key,
            project_id integer not null,
            source text,
            provenance_json text
        );
        create table if not exists project_metadata (
            id integer primary key,
            project_id integer not null,
            key text not null,
            value text,
            provenance_json text
        );
        create table if not exists result_provenance (
            id integer primary key,
            project_id integer not null,
            result_table text not null,
            result_key text not null,
            field_name text,
            source_handles text,
            method text not null,
            configuration_hash text not null,
            confidence real not null,
            warning text
        );
        """
    )


def _read_rows(path: Path) -> list[tuple[int, dict[str, str | None]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [(index, _normalize_row(row)) for index, row in enumerate(csv.DictReader(handle), start=2)]
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, ())
        keys = [_normalize_key(value) for value in headers]
        return [
            (index, _normalize_row(dict(zip(keys, values, strict=False))))
            for index, values in enumerate(rows, start=2)
            if any(value is not None for value in values)
        ]
    raise ValueError(f"unsupported metadata file: {path}")


def _normalize_row(row: dict[str, object]) -> dict[str, str | None]:
    return {_normalize_key(key): _normalize_value(value) for key, value in row.items() if _normalize_key(key)}


def _normalize_key(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_value(value: object) -> str | None:
    text = str(value or "").strip()
    return None if text.lower() in UNKNOWN_VALUES else text


def _resolve_row_project(session: Session, row: dict[str, str | None]) -> Project | None:
    drawing_id = row.get("drawing_id")
    if drawing_id:
        project = session.scalar(select(Project).where(Project.drawing_id == drawing_id))
        if project is not None:
            return project
    code = row.get("project_code")
    if not code:
        return None
    code = code.upper()
    mapped = session.scalar(select(ProjectCode).where(ProjectCode.code == code))
    if mapped is not None:
        return session.get(Project, mapped.project_id)
    matches = [
        project
        for project in session.scalars(select(Project)).all()
        if code in resolve_project_codes(project.drawing_id).related_project_codes
    ]
    return matches[0] if len(matches) == 1 else None


def _resolve_sqlite_project(db: sqlite3.Connection, row: dict[str, str | None]) -> tuple[int, str] | None:
    drawing_id = row.get("drawing_id")
    if drawing_id:
        project = db.execute("select id, drawing_id from projects where drawing_id = ?", (drawing_id,)).fetchone()
        if project is not None:
            return int(project[0]), str(project[1])
    code = row.get("project_code")
    if not code:
        return None
    code = code.upper()
    project = db.execute(
        "select p.id, p.drawing_id from project_codes pc join projects p on p.id = pc.project_id where pc.code = ?",
        (code,),
    ).fetchone()
    if project is not None:
        return int(project[0]), str(project[1])
    matches = [
        (int(project_id), str(project_drawing_id))
        for project_id, project_drawing_id in db.execute("select id, drawing_id from projects")
        if code in resolve_project_codes(str(project_drawing_id)).related_project_codes
    ]
    return matches[0] if len(matches) == 1 else None


def _upsert_metadata(
    session: Session,
    project_id: int,
    path: Path,
    row_number: int,
    row: dict[str, str | None],
) -> list[str]:
    conflicts: list[str] = []
    for field in APPROVED_FIELDS:
        if field not in row:
            continue
        value = row[field]
        existing = session.scalar(
            select(ProjectMetadata).where(ProjectMetadata.project_id == project_id).where(ProjectMetadata.key == field)
        )
        if existing is not None and existing.value != value:
            conflicts.append(field)
            _add_provenance(session, project_id, field, path, row_number, warning="conflicting_metadata")
            continue
        if existing is None:
            existing = ProjectMetadata(project_id=project_id, key=field, value=value, provenance_json={})
            session.add(existing)
        existing.value = value
        existing.provenance_json = {"source_file": str(path), "row": row_number}
        _add_provenance(session, project_id, field, path, row_number)
    return conflicts


def _upsert_sqlite_metadata(
    db: sqlite3.Connection,
    project_id: int,
    path: Path,
    row_number: int,
    row: dict[str, str | None],
) -> list[str]:
    conflicts: list[str] = []
    for field in APPROVED_FIELDS:
        if field not in row:
            continue
        value = row[field]
        existing = db.execute(
            "select value from project_metadata where project_id = ? and key = ?",
            (project_id, field),
        ).fetchone()
        if existing is not None and existing[0] != value:
            conflicts.append(field)
            _add_sqlite_provenance(db, project_id, field, path, row_number, warning="conflicting_metadata")
            continue
        provenance = json.dumps({"source_file": str(path), "row": row_number}, sort_keys=True)
        if existing is None:
            db.execute(
                "insert into project_metadata (project_id, key, value, provenance_json) values (?, ?, ?, ?)",
                (project_id, field, value, provenance),
            )
        else:
            db.execute(
                "update project_metadata set value = ?, provenance_json = ? where project_id = ? and key = ?",
                (value, provenance, project_id, field),
            )
        _add_sqlite_provenance(db, project_id, field, path, row_number)
    return conflicts


def _ensure_project_codes(session: Session, project: Project, row: dict[str, str | None]) -> None:
    codes = set(resolve_project_codes(project.drawing_id).related_project_codes)
    if row.get("project_code"):
        codes.add(row["project_code"].upper())
    for code in codes:
        if session.get(ProjectCode, code) is None:
            session.add(ProjectCode(code=code, project_id=project.id, source="metadata_import", provenance_json={}))


def _ensure_sqlite_project_codes(
    db: sqlite3.Connection,
    project_id: int,
    drawing_id: str,
    row: dict[str, str | None],
) -> None:
    codes = set(resolve_project_codes(drawing_id).related_project_codes)
    if row.get("project_code"):
        codes.add(row["project_code"].upper())
    for code in codes:
        db.execute(
            "insert or ignore into project_codes (code, project_id, source, provenance_json) values (?, ?, ?, ?)",
            (code, project_id, "metadata_import", "{}"),
        )


def _add_provenance(
    session: Session,
    project_id: int,
    field: str,
    path: Path,
    row_number: int,
    *,
    warning: str | None = None,
) -> None:
    session.add(
        ResultProvenance(
            project_id=project_id,
            result_table="project_metadata",
            result_key="metadata",
            field_name=field,
            source_handles=[f"{path}:{row_number}"],
            method="metadata_import",
            configuration_hash="metadata_import",
            confidence=0.5 if warning else 1.0,
            warning=warning,
        )
    )


def _add_sqlite_provenance(
    db: sqlite3.Connection,
    project_id: int,
    field: str,
    path: Path,
    row_number: int,
    *,
    warning: str | None = None,
) -> None:
    db.execute(
        "insert into result_provenance "
        "(project_id, result_table, result_key, field_name, source_handles, method, configuration_hash, confidence, warning) "
        "values (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            project_id,
            "project_metadata",
            "metadata",
            field,
            json.dumps([f"{path}:{row_number}"]),
            "metadata_import",
            "metadata_import",
            0.5 if warning else 1.0,
            warning,
        ),
    )
